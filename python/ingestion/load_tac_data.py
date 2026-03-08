"""
load_tac_data.py
----------------
Ingests NHS England TAC (Trust Accounts Consolidation) Excel files
into MySQL databases nhs_stg (staging) and nhs_finance (fact table).

Usage:
    python python/ingestion/load_tac_data.py

Expects files in data/raw/ matching:
    TAC_NHS_trusts_YYYY-YY.xlsx
    TAC_NHS_foundation_trusts_YYYY-YY.xlsx
"""

import logging
import re
from pathlib import Path

import pandas as pd
from sqlalchemy import create_engine, text

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Config ─────────────────────────────────────────────────────────────────

DB_USER     = "root"
DB_PASSWORD = "Password1234"
DB_HOST     = "127.0.0.1"
DB_PORT     = 3306

STG_DB      = "nhs_stg"
FACT_DB     = "nhs_finance"

RAW_DIR       = Path("data/raw")
PROCESSED_DIR = Path("data/processed")

YEAR_RE = re.compile(r"(\d{4})-(\d{2})")

RAW_COLUMNS = [
    "OrganisationName", "WorkSheetName", "TableID",
    "MainCode", "RowNumber", "SubCode", "Total",
]

# ── Helpers ────────────────────────────────────────────────────────────────

def get_engine(database: str):
    url = f"mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{database}?charset=utf8mb4"
    return create_engine(url, echo=False)


def filename_to_financial_year(filename: str) -> str:
    match = YEAR_RE.search(filename)
    if not match:
        raise ValueError(f"Cannot extract financial year from: {filename}")
    return f"{match.group(1)}/{match.group(2)}"


def filename_to_trust_type(filename: str) -> str:
    return "FOUNDATION_TRUST" if "foundation" in filename.lower() else "NHS_TRUST"


def infer_year_type(main_code: str) -> str:
    if "PY" in str(main_code):
        return "PY"
    return "CY"


# ── Readers ────────────────────────────────────────────────────────────────

def read_provider_list(path: Path, source_file: str, financial_year: str) -> pd.DataFrame:
    log.info("  Reading provider list...")
    df = pd.read_excel(path, sheet_name="List of Providers", header=0)
    df = df.rename(columns={
        "Full name of Provider": "organisation_name",
        "NHS code":              "org_code",
        "Region":                "region",
        "Sector":                "sector",
        "Comments ":             "comments",
    })
    # Handle column name with/without trailing space
    if "Comments" in df.columns and "comments" not in df.columns:
        df = df.rename(columns={"Comments": "comments"})

    keep = ["organisation_name", "org_code", "region", "sector", "comments"]
    df = df[[c for c in keep if c in df.columns]].copy()
    df["org_code"]          = df["org_code"].astype(str).str.strip()
    df["organisation_name"] = df["organisation_name"].astype(str).str.strip()
    df["source_file"]       = source_file
    df["financial_year"]    = financial_year
    df = df.dropna(subset=["org_code"])
    df = df[df["org_code"].str.len() == 3]
    log.info("  Found %d providers", len(df))
    return df


def find_sheet_name(path: Path, target: str) -> str:
    """Return the actual sheet name matching target case-insensitively."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    for name in wb.sheetnames:
        if name.lower() == target.lower():
            return name
    raise ValueError(f"Sheet '{target}' not found in {path.name}. Available: {wb.sheetnames}")


def read_all_data(path: Path, source_file: str, financial_year: str, trust_type: str) -> pd.DataFrame:
    log.info("  Reading All data sheet (this takes ~30s for large files)...")
    sheet_name = find_sheet_name(path, "All data")
    df = pd.read_excel(
        path,
        sheet_name=sheet_name,
        header=0,
        dtype={
            "OrganisationName": str,
            "WorkSheetName":    str,
            "TableID":          "Int64",
            "MainCode":         str,
            "RowNumber":        "Int64",
            "SubCode":          str,
            "Total":            float,
        },
    )

    # Normalise column names across file versions
    # 2023/24: OrganisationName, Total
    # 2021/22: Organisation Name, Value number
    col_map = {
        "Organisation Name": "OrganisationName",
        "Value number":      "Total",
        "Value Number":      "Total",
    }
    df = df.rename(columns=col_map)

    missing = [c for c in RAW_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns in {path.name}: {missing}. Found: {list(df.columns)}")

    df = df[RAW_COLUMNS].copy()
    df.columns = ["organisation_name", "worksheet_name", "table_id",
                  "main_code", "row_num", "sub_code", "total"]

    df["organisation_name"] = df["organisation_name"].astype(str).str.strip()
    df["worksheet_name"]    = df["worksheet_name"].astype(str).str.strip()
    df["main_code"]         = df["main_code"].astype(str).str.strip()
    df["sub_code"]          = df["sub_code"].astype(str).str.strip()

    before = len(df)
    df = df.dropna(subset=["organisation_name", "sub_code", "total"])
    dropped = before - len(df)
    if dropped:
        log.warning("  Dropped %d null rows", dropped)

    # Add year_type from MainCode, then keep CY only
    df["year_type"] = df["main_code"].apply(infer_year_type)
    df = df[df["year_type"] == "CY"].copy()

    df["total"]          = pd.to_numeric(df["total"], errors="coerce").fillna(0).astype("int64")
    df["source_file"]    = source_file
    df["trust_type"]     = trust_type
    df["financial_year"] = financial_year

    log.info("  Loaded %d CY rows across %d providers",
             len(df), df["organisation_name"].nunique())
    return df


# ── Validation ─────────────────────────────────────────────────────────────

def validate(data_df: pd.DataFrame, provider_df: pd.DataFrame, filename: str) -> None:
    errors = []
    for col in ["organisation_name", "sub_code", "total", "financial_year"]:
        nulls = data_df[col].isna().sum()
        if nulls:
            errors.append(f"[CRITICAL] {nulls} nulls in '{col}'")

    dupes = data_df.duplicated(subset=["organisation_name", "main_code", "sub_code"]).sum()
    if dupes:
        log.warning("  %d duplicate (org, main_code, sub_code) rows in %s", dupes, filename)

    unmatched = set(data_df["organisation_name"]) - set(provider_df["organisation_name"])
    if unmatched:
        log.warning("  %d org names not in provider list: %s", len(unmatched), list(unmatched)[:3])

    if errors:
        for e in errors:
            log.error(e)
        raise ValueError(f"Critical validation failures in {filename}")

    log.info("  Validation passed")


# ── Database loaders ───────────────────────────────────────────────────────

def load_staging(data_df: pd.DataFrame, provider_df: pd.DataFrame,
                 financial_year: str, trust_type: str) -> None:
    engine = get_engine(STG_DB)

    with engine.begin() as conn:
        conn.execute(text(
            "DELETE FROM stg_tac_raw WHERE financial_year = :fy AND trust_type = :tt"
        ), {"fy": financial_year, "tt": trust_type})
        conn.execute(text(
            "DELETE FROM stg_provider_list WHERE financial_year = :fy AND trust_type = :tt"
        ), {"fy": financial_year, "tt": trust_type})

    provider_df["trust_type"] = trust_type
    provider_df.to_sql("stg_provider_list", engine, if_exists="append", index=False, chunksize=500)
    log.info("  Loaded %d rows → nhs_stg.stg_provider_list", len(provider_df))

    data_df.to_sql("stg_tac_raw", engine, if_exists="append", index=False, chunksize=2000)
    log.info("  Loaded %d rows → nhs_stg.stg_tac_raw", len(data_df))


def populate_dim_trust(financial_year: str) -> None:
    stg_engine  = get_engine(STG_DB)
    fact_engine = get_engine(FACT_DB)

    providers = pd.read_sql(
        "SELECT DISTINCT org_code, organisation_name, sector, region, source_file, financial_year "
        "FROM stg_provider_list WHERE LENGTH(TRIM(org_code)) = 3",
        stg_engine,
    )
    if providers.empty:
        log.warning("  No providers found in staging — skipping dim_trust")
        return

    providers["trust_type"]   = providers["source_file"].apply(
        lambda f: "FOUNDATION_TRUST" if "foundation" in f.lower() else "NHS_TRUST"
    )
    providers["is_foundation"] = (providers["trust_type"] == "FOUNDATION_TRUST").astype(int)

    upsert_sql = text("""
        INSERT INTO dim_trust
            (org_code, organisation_name, trust_type, sector, region, is_foundation,
             first_year_seen, last_year_seen)
        VALUES
            (:org_code, :organisation_name, :trust_type, :sector, :region, :is_foundation,
             :financial_year, :financial_year)
        ON DUPLICATE KEY UPDATE
            last_year_seen   = VALUES(last_year_seen),
            sector           = COALESCE(VALUES(sector), sector),
            region           = COALESCE(VALUES(region), region),
            updated_ts       = CURRENT_TIMESTAMP
    """)

    with fact_engine.begin() as conn:
        for _, row in providers.iterrows():
            conn.execute(upsert_sql, {
                "org_code":          row["org_code"],
                "organisation_name": row["organisation_name"],
                "trust_type":        row["trust_type"],
                "sector":            row.get("sector"),
                "region":            row.get("region"),
                "is_foundation":     int(row["is_foundation"]),
                "financial_year":    financial_year,
            })

    log.info("  Upserted %d rows → nhs_finance.dim_trust", len(providers))


def promote_to_fact(financial_year: str, trust_type: str) -> None:
    """Join staging data with provider list to get org_code, then upsert into fct_tac."""
    stg_engine  = get_engine(STG_DB)
    fact_engine = get_engine(FACT_DB)

    log.info("  Promoting to fct_tac (year=%s, type=%s)...", financial_year, trust_type)

    # Ensure all worksheet names from this file exist in dim_worksheet
    stg_sheets = pd.read_sql(
        "SELECT DISTINCT worksheet_name FROM stg_tac_raw WHERE financial_year = %(fy)s AND trust_type = %(tt)s",
        stg_engine, params={"fy": financial_year, "tt": trust_type}
    )
    with fact_engine.begin() as conn:
        for ws_name in stg_sheets["worksheet_name"]:
            conn.execute(text("""
                INSERT IGNORE INTO dim_worksheet (worksheet_name, schedule_title, category)
                VALUES (:ws, :ws, 'OTHER')
            """), {"ws": ws_name})

    # Load staging data joined with org_code
    sql = """
        SELECT
            p.org_code,
            r.financial_year,
            r.worksheet_name,
            r.table_id,
            r.main_code,
            r.sub_code,
            r.total          AS total_000s,
            r.trust_type,
            r.source_file
        FROM stg_tac_raw r
        JOIN stg_provider_list p
            ON  r.organisation_name = p.organisation_name
            AND r.financial_year    = p.financial_year
            AND r.trust_type        = p.trust_type
        WHERE r.financial_year = %(fy)s
          AND r.trust_type     = %(tt)s
    """
    fact_df = pd.read_sql(sql, stg_engine, params={"fy": financial_year, "tt": trust_type})
    log.info("  Joined %d rows ready for fact table", len(fact_df))

    if fact_df.empty:
        log.warning("  No rows to promote — check staging data")
        return

    # Upsert in chunks
    upsert_sql = text("""
        INSERT INTO fct_tac
            (org_code, financial_year, worksheet_name, table_id,
             main_code, sub_code, total_000s, trust_type, source_file)
        VALUES
            (:org_code, :financial_year, :worksheet_name, :table_id,
             :main_code, :sub_code, :total_000s, :trust_type, :source_file)
        ON DUPLICATE KEY UPDATE
            total_000s  = VALUES(total_000s),
            source_file = VALUES(source_file),
            load_ts     = CURRENT_TIMESTAMP
    """)

    chunk_size = 1000
    total_rows = 0
    with fact_engine.begin() as conn:
        for start in range(0, len(fact_df), chunk_size):
            chunk = fact_df.iloc[start: start + chunk_size]
            conn.execute(upsert_sql, chunk.to_dict(orient="records"))
            total_rows += len(chunk)

    log.info("  Upserted %d rows → nhs_finance.fct_tac", total_rows)


# ── Main ───────────────────────────────────────────────────────────────────

def process_file(path: Path) -> None:
    source_file    = path.name
    financial_year = filename_to_financial_year(source_file)
    trust_type     = filename_to_trust_type(source_file)

    log.info("=" * 60)
    log.info("FILE:  %s", source_file)
    log.info("YEAR:  %s  |  TYPE: %s", financial_year, trust_type)
    log.info("=" * 60)

    provider_df = read_provider_list(path, source_file, financial_year)
    data_df     = read_all_data(path, source_file, financial_year, trust_type)

    validate(data_df, provider_df, source_file)
    load_staging(data_df, provider_df, financial_year, trust_type)
    populate_dim_trust(financial_year)
    promote_to_fact(financial_year, trust_type)

    log.info("  DONE: %s\n", source_file)


def main():
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    files = sorted(
        p for p in RAW_DIR.glob("TAC_NHS_*.xlsx")
        if "illustrative" not in p.name.lower()
    )

    if not files:
        log.error("No TAC data files found in %s", RAW_DIR)
        return

    log.info("Found %d files to process", len(files))
    for path in files:
        try:
            process_file(path)
        except Exception as exc:
            log.error("FAILED: %s — %s", path.name, exc)
            raise

    # Final counts
    engine = get_engine(FACT_DB)
    with engine.connect() as conn:
        facts  = conn.execute(text("SELECT COUNT(*) FROM fct_tac")).scalar()
        trusts = conn.execute(text("SELECT COUNT(*) FROM dim_trust")).scalar()

    log.info("=" * 60)
    log.info("COMPLETE")
    log.info("  dim_trust rows : %d", trusts)
    log.info("  fct_tac rows   : %d", facts)
    log.info("=" * 60)


if __name__ == "__main__":
    main()
